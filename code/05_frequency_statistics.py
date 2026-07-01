from pathlib import Path
import re
import warnings
import unicodedata

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from scipy import stats
import statsmodels.formula.api as smf
from statsmodels.stats.multitest import multipletests

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REPO_DIR = Path(__file__).resolve().parents[1]
BASE_DIR = REPO_DIR / "outputs"

INPUT_FILE = BASE_DIR / "frequency_results.xlsx"
INPUT_SHEET = "frequency_results"

OUTPUT_FILE = BASE_DIR / "frequency_statistics_results.xlsx"
PLOTS_DIR = BASE_DIR / "frequency_plots"

PRIMARY_VAR = "log_mnf_10cycles_ratio"
PLOT_VAR = "mnf_10cycles_ratio"

TRIAL_ORDER = [
    "Pre",
    "Post",
    "Post5",
    "Post10",
    "Post15",
    "Post20",
    "Post25",
    "Post30",
]

TRIAL_LABELS = {
    "Pre": "Pre",
    "Post": "Post",
    "Post5": "Post 5",
    "Post10": "Post 10",
    "Post15": "Post 15",
    "Post20": "Post 20",
    "Post25": "Post 25",
    "Post30": "Post 30",
}

INTERGROUP_TRIALS = [
    "Post",
    "Post5",
    "Post10",
    "Post15",
    "Post20",
    "Post25",
    "Post30",
]

MUSCLE_ORDER = ["biceps", "triceps"]
SESSION_ORDER = ["control", "vibration"]

SESSION_LABELS = {
    "control": "Control",
    "vibration": "Vibration",
}

MUSCLE_LABELS = {
    "biceps": "Biceps",
    "triceps": "Triceps",
}

MUSCLE_PLOT_LABELS = {
    "biceps": "Biceps brachii",
    "triceps": "Triceps brachii",
}


def normalize_text(value):
    if pd.isna(value):
        return ""

    text = str(value).strip().lower()

    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))

    return text



def prepare_data(df):
    df = df.copy()

    required_columns = [
        "subject_id",
        "session",
        "trial_cat",
        "muscle",
        "feature_status",
        "include_analysis",
        "qc_flag",
        "mnf_10cycles",
        "mnf_10cycles_ratio",
        "log_mnf_10cycles_ratio",
    ]

    missing_columns = [
        col for col in required_columns
        if col not in df.columns
    ]

    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")

    df["session_norm"] = df["session"].apply(normalize_text)
    df["muscle_norm"] = df["muscle"].apply(normalize_text)

    df["trial_cat"] = pd.Categorical(
        df["trial_cat"],
        categories=TRIAL_ORDER,
        ordered=True,
    )

    df["feature_status_norm"] = df["feature_status"].apply(normalize_text)
    df["qc_flag_norm"] = df["qc_flag"].apply(normalize_text)

    df["include_analysis_bool"] = (
        df["include_analysis"]
        .astype(str)
        .str.lower()
        .isin(["true", "1", "yes", "si", "sí"])
    )

    for col in [
        "mnf_10cycles",
        "mnf_10cycles_ratio",
        "log_mnf_10cycles_ratio",
        "line_noise_ratio_49_51",
        "peak_frequency_no50_10cycles",
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")


    return df


def apply_bh_correction(df, p_col="p_value", group_cols=None):
    df = df.copy()

    df["p_value_bh"] = np.nan
    df["significant_bh"] = False

    if df.empty or p_col not in df.columns:
        return df

    if group_cols is None:
        group_cols = []

    if not group_cols:
        valid = df[p_col].notna()

        if valid.sum() > 0:
            _, p_bh, _, _ = multipletests(
                df.loc[valid, p_col],
                alpha=0.05,
                method="fdr_bh",
            )

            df.loc[valid, "p_value_bh"] = p_bh
            df.loc[valid, "significant_bh"] = p_bh < 0.05

        return df

    for _, group in df.groupby(group_cols, dropna=False):
        valid_idx = group.index[group[p_col].notna()]

        if len(valid_idx) == 0:
            continue

        _, p_bh, _, _ = multipletests(
            df.loc[valid_idx, p_col],
            alpha=0.05,
            method="fdr_bh",
        )

        df.loc[valid_idx, "p_value_bh"] = p_bh
        df.loc[valid_idx, "significant_bh"] = p_bh < 0.05

    return df


def parse_trial_from_lmm_term(term):
    match = re.search(r"\[T\.(.*?)\]", term)

    if match:
        return match.group(1)

    return term


def clean_float(value, decimals=4):
    if pd.isna(value) or not np.isfinite(value):
        return np.nan

    return round(float(value), decimals)


def summarize_mnf(df, dataset_name):
    rows = []

    for muscle in MUSCLE_ORDER:
        for session in SESSION_ORDER:
            for trial in TRIAL_ORDER:
                subset = df[
                    (df["muscle_norm"] == muscle)
                    & (df["session_norm"] == session)
                    & (df["trial_cat"] == trial)
                    & (df[PLOT_VAR].notna())
                ].copy()

                values = subset[PLOT_VAR].dropna()

                if values.empty:
                    continue

                rows.append(
                    {
                        "dataset": dataset_name,
                        "muscle": MUSCLE_LABELS.get(muscle, muscle),
                        "session": SESSION_LABELS.get(session, session),
                        "trial": trial,
                        "n": int(values.count()),
                        "mean_mnf_ratio": clean_float(values.mean()),
                        "sd_mnf_ratio": clean_float(values.std(ddof=1)),
                        "sem_mnf_ratio": clean_float(values.sem()),
                        "median_mnf_ratio": clean_float(values.median()),
                        "q1_mnf_ratio": clean_float(values.quantile(0.25)),
                        "q3_mnf_ratio": clean_float(values.quantile(0.75)),
                        "min_mnf_ratio": clean_float(values.min()),
                        "max_mnf_ratio": clean_float(values.max()),
                        "mean_mnf_hz": clean_float(subset["mnf_10cycles"].mean()),
                        "median_mnf_hz": clean_float(subset["mnf_10cycles"].median()),
                    }
                )

    return pd.DataFrame(rows)


def run_intragroup_lmm(df, dataset_name):
    rows = []
    model_notes = []

    post_trials = INTERGROUP_TRIALS
    reference_trial = "Post"

    for muscle in MUSCLE_ORDER:
        for session in SESSION_ORDER:
            subset = df[
                (df["muscle_norm"] == muscle)
                & (df["session_norm"] == session)
                & (df["trial_cat"].isin(post_trials))
                & (df[PRIMARY_VAR].notna())
                & np.isfinite(df[PRIMARY_VAR])
            ].copy()

            n_subjects = subset["subject_id"].nunique()
            n_observations = len(subset)

            if subset.empty:
                for trial in post_trials:
                    rows.append(
                        {
                            "dataset": dataset_name,
                            "muscle": MUSCLE_LABELS.get(muscle, muscle),
                            "session": SESSION_LABELS.get(session, session),
                            "comparison": f"{trial} vs Pre",
                            "estimate_log": np.nan,
                            "estimated_ratio_vs_pre": np.nan,
                            "percent_change_vs_pre": np.nan,
                            "std_error": np.nan,
                            "z_value": np.nan,
                            "p_value": np.nan,
                            "n_subjects": 0,
                            "n_observations": 0,
                            "model_status": "no_data",
                        }
                    )
                continue

            if n_subjects < 3:
                for trial in post_trials:
                    rows.append(
                        {
                            "dataset": dataset_name,
                            "muscle": MUSCLE_LABELS.get(muscle, muscle),
                            "session": SESSION_LABELS.get(session, session),
                            "comparison": f"{trial} vs Pre",
                            "estimate_log": np.nan,
                            "estimated_ratio_vs_pre": np.nan,
                            "percent_change_vs_pre": np.nan,
                            "std_error": np.nan,
                            "z_value": np.nan,
                            "p_value": np.nan,
                            "n_subjects": n_subjects,
                            "n_observations": n_observations,
                            "model_status": "not_enough_subjects",
                        }
                    )
                continue

            subset["trial_cat"] = subset["trial_cat"].astype(str)

            subset["trial_cat"] = pd.Categorical(
                subset["trial_cat"],
                categories=post_trials,
                ordered=True,
            )

            subset["subject_id"] = subset["subject_id"].astype(str)

            formula = f"{PRIMARY_VAR} ~ C(trial_cat, Treatment(reference='{reference_trial}'))"

            try:
                result, model_status, warning_text = _fit_mixedlm_with_fallback(
                    formula=formula,
                    data=subset,
                    groups_col="subject_id",
                )

                if warning_text:
                    model_notes.append(
                        {
                            "dataset": dataset_name,
                            "analysis": "intragroup_lmm",
                            "muscle": MUSCLE_LABELS.get(muscle, muscle),
                            "session": SESSION_LABELS.get(session, session),
                            "note": warning_text,
                        }
                    )

                fe_names = list(result.fe_params.index)
                cov = result.cov_params().loc[fe_names, fe_names].to_numpy(dtype=float)

                for trial in post_trials:
                    contrast = np.zeros(len(fe_names))

                    if trial == reference_trial:
                        if "Intercept" not in fe_names:
                            rows.append(
                                {
                                    "dataset": dataset_name,
                                    "muscle": MUSCLE_LABELS.get(muscle, muscle),
                                    "session": SESSION_LABELS.get(session, session),
                                    "comparison": f"{trial} vs Pre",
                                    "estimate_log": np.nan,
                                    "estimated_ratio_vs_pre": np.nan,
                                    "percent_change_vs_pre": np.nan,
                                    "std_error": np.nan,
                                    "z_value": np.nan,
                                    "p_value": np.nan,
                                    "n_subjects": n_subjects,
                                    "n_observations": n_observations,
                                    "model_status": "contrast_not_estimable_missing_intercept",
                                }
                            )
                            continue

                        contrast[fe_names.index("Intercept")] = 1.0

                    else:
                        term = (
                            f"C(trial_cat, Treatment(reference='{reference_trial}'))"
                            f"[T.{trial}]"
                        )

                        if "Intercept" not in fe_names or term not in fe_names:
                            rows.append(
                                {
                                    "dataset": dataset_name,
                                    "muscle": MUSCLE_LABELS.get(muscle, muscle),
                                    "session": SESSION_LABELS.get(session, session),
                                    "comparison": f"{trial} vs Pre",
                                    "estimate_log": np.nan,
                                    "estimated_ratio_vs_pre": np.nan,
                                    "percent_change_vs_pre": np.nan,
                                    "std_error": np.nan,
                                    "z_value": np.nan,
                                    "p_value": np.nan,
                                    "n_subjects": n_subjects,
                                    "n_observations": n_observations,
                                    "model_status": "contrast_not_estimable_missing_trial",
                                }
                            )
                            continue

                        contrast[fe_names.index("Intercept")] = 1.0
                        contrast[fe_names.index(term)] = 1.0

                    estimate = float(np.dot(contrast, result.fe_params.loc[fe_names].values))
                    variance = float(np.dot(contrast, np.dot(cov, contrast)))

                    if variance < 0 and np.isclose(variance, 0):
                        variance = 0.0

                    if variance >= 0 and np.isfinite(variance):
                        std_error = float(np.sqrt(variance))
                    else:
                        std_error = np.nan

                    if pd.notna(std_error) and std_error > 0:
                        z_value = estimate / std_error
                        p_value = float(2 * stats.norm.sf(abs(z_value)))
                    else:
                        z_value = np.nan
                        p_value = np.nan

                    estimated_ratio = np.exp(estimate)
                    percent_change = (estimated_ratio - 1) * 100

                    rows.append(
                        {
                            "dataset": dataset_name,
                            "muscle": MUSCLE_LABELS.get(muscle, muscle),
                            "session": SESSION_LABELS.get(session, session),
                            "comparison": f"{trial} vs Pre",
                            "estimate_log": clean_float(estimate),
                            "estimated_ratio_vs_pre": clean_float(estimated_ratio),
                            "percent_change_vs_pre": clean_float(percent_change),
                            "std_error": clean_float(std_error),
                            "z_value": clean_float(z_value),
                            "p_value": p_value,
                            "n_subjects": n_subjects,
                            "n_observations": n_observations,
                            "model_status": model_status,
                        }
                    )

            except Exception as exc:
                error_text = str(exc)

                if "singular" in error_text.lower():
                    model_status = "not_estimated_singular_matrix"
                    note_text = "Model not estimated due to singular matrix."
                else:
                    model_status = f"model_error: {error_text}"
                    note_text = error_text

                for trial in post_trials:
                    rows.append(
                        {
                            "dataset": dataset_name,
                            "muscle": MUSCLE_LABELS.get(muscle, muscle),
                            "session": SESSION_LABELS.get(session, session),
                            "comparison": f"{trial} vs Pre",
                            "estimate_log": np.nan,
                            "estimated_ratio_vs_pre": np.nan,
                            "percent_change_vs_pre": np.nan,
                            "std_error": np.nan,
                            "z_value": np.nan,
                            "p_value": np.nan,
                            "n_subjects": n_subjects,
                            "n_observations": n_observations,
                            "model_status": model_status,
                        }
                    )

                model_notes.append(
                    {
                        "dataset": dataset_name,
                        "analysis": "intragroup_lmm",
                        "muscle": MUSCLE_LABELS.get(muscle, muscle),
                        "session": SESSION_LABELS.get(session, session),
                        "note": note_text,
                    }
                )

    results = pd.DataFrame(rows)

    if not results.empty:
        results = apply_bh_correction(
            results,
            p_col="p_value",
            group_cols=["dataset", "muscle", "session"],
        )

    notes = pd.DataFrame(model_notes)

    return results, notes

def _fit_mixedlm_with_fallback(formula, data, groups_col="subject_id"):
    optimizers = ["lbfgs", "bfgs", "cg", "powell", "nm"]
    fit_attempts = []

    for method in optimizers:
        try:
            with warnings.catch_warnings(record=True) as caught_warnings:
                warnings.simplefilter("always")

                model = smf.mixedlm(
                    formula,
                    data=data,
                    groups=data[groups_col],
                )

                result = model.fit(
                    reml=False,
                    method=method,
                    maxiter=500,
                    disp=False,
                )

                warning_text = " | ".join(str(w.message) for w in caught_warnings)

            converged = bool(getattr(result, "converged", False))

            if converged:
                if warning_text:
                    status = f"ok_{method}_with_warning"
                else:
                    status = f"ok_{method}"

                return result, status, warning_text

            fit_attempts.append(
                f"{method}: not_converged"
                + (f"; warnings: {warning_text}" if warning_text else "")
            )

        except Exception as exc:
            fit_attempts.append(f"{method}: error: {exc}")

    raise RuntimeError("No optimizer converged. Attempts: " + " || ".join(fit_attempts))


def _mean_ratio_for_cell(df, session, trial):
    values = df.loc[
        (df["session_norm"] == session)
        & (df["trial_cat"] == trial),
        PLOT_VAR,
    ].dropna()

    if values.empty:
        return np.nan

    return clean_float(values.mean())


def _paired_subject_count_for_trial(df, trial):
    trial_df = df[df["trial_cat"] == trial].copy()

    control_subjects = set(
        trial_df.loc[trial_df["session_norm"] == "control", "subject_id"].dropna()
    )
    vibration_subjects = set(
        trial_df.loc[trial_df["session_norm"] == "vibration", "subject_id"].dropna()
    )

    return len(control_subjects.intersection(vibration_subjects))


def _empty_intergroup_lmm_row(
    dataset_name,
    muscle,
    trial,
    model_status,
    n_subjects=0,
    n_observations=0,
    n_subjects_contrast=0,
    control_mean_ratio=np.nan,
    vibration_mean_ratio=np.nan,
):
    model_name = f"{PRIMARY_VAR} ~ 0 + C(cell) + (1 | subject_id)"

    return {
        "dataset": dataset_name,
        "muscle": MUSCLE_LABELS.get(muscle, muscle),
        "trial": trial,
        "time_point": trial,
        "analysis_type": "intergroup_lmm",
        "model_name": model_name,
        "comparison": f"vibration_{trial} - control_{trial}",
        "contrast": f"C(cell)[vibration_{trial}] - C(cell)[control_{trial}]",
        "estimate_log_difference": np.nan,
        "mean_log_difference_vibration_minus_control": np.nan,
        "standard_error": np.nan,
        "std_error": np.nan,
        "z_value": np.nan,
        "p_raw": np.nan,
        "p_value": np.nan,
        "p_adj": np.nan,
        "p_value_bh": np.nan,
        "significant_raw": False,
        "significant_adj": False,
        "significant_bh": False,
        "ratio_of_ratios": np.nan,
        "percent_difference_vibration_vs_control": np.nan,
        "ci_low_log": np.nan,
        "ci_high_log": np.nan,
        "ci_low_ratio": np.nan,
        "ci_high_ratio": np.nan,
        "control_mean_mnf_ratio": control_mean_ratio,
        "vibration_mean_mnf_ratio": vibration_mean_ratio,
        "n_subjects": int(n_subjects),
        "n_observations": int(n_observations),
        "n_subjects_contrast": int(n_subjects_contrast),
        "model_status": model_status,
        "convergence_status": model_status,
    }


def run_intergroup_tests(df, dataset_name):
    rows = []
    z_crit = stats.norm.ppf(0.975)
    model_name = f"{PRIMARY_VAR} ~ 0 + C(cell) + (1 | subject_id)"

    for muscle in MUSCLE_ORDER:
        muscle_df = df[
            (df["muscle_norm"] == muscle)
            & (df["session_norm"].isin(SESSION_ORDER))
            & (df["trial_cat"].isin(INTERGROUP_TRIALS))
            & (df[PRIMARY_VAR].notna())
            & np.isfinite(df[PRIMARY_VAR])
        ].copy()

        n_subjects = muscle_df["subject_id"].nunique()
        n_observations = len(muscle_df)

        if muscle_df.empty:
            for trial in INTERGROUP_TRIALS:
                rows.append(
                    _empty_intergroup_lmm_row(
                        dataset_name=dataset_name,
                        muscle=muscle,
                        trial=trial,
                        model_status="no_data_for_muscle",
                    )
                )
            continue

        muscle_df["trial_name"] = muscle_df["trial_cat"].astype(str)
        muscle_df["cell"] = muscle_df["session_norm"] + "_" + muscle_df["trial_name"]
        muscle_df["subject_id"] = muscle_df["subject_id"].astype(str)

        if n_subjects < 3:
            for trial in INTERGROUP_TRIALS:
                rows.append(
                    _empty_intergroup_lmm_row(
                        dataset_name=dataset_name,
                        muscle=muscle,
                        trial=trial,
                        model_status="not_enough_subjects_for_lmm",
                        n_subjects=n_subjects,
                        n_observations=n_observations,
                        n_subjects_contrast=_paired_subject_count_for_trial(muscle_df, trial),
                        control_mean_ratio=_mean_ratio_for_cell(muscle_df, "control", trial),
                        vibration_mean_ratio=_mean_ratio_for_cell(muscle_df, "vibration", trial),
                    )
                )
            continue

        observed_cells = set(muscle_df["cell"].dropna().unique())
        has_any_valid_contrast = any(
            f"control_{trial}" in observed_cells
            and f"vibration_{trial}" in observed_cells
            for trial in INTERGROUP_TRIALS
        )

        if not has_any_valid_contrast:
            for trial in INTERGROUP_TRIALS:
                rows.append(
                    _empty_intergroup_lmm_row(
                        dataset_name=dataset_name,
                        muscle=muscle,
                        trial=trial,
                        model_status="no_estimable_control_vibration_contrasts",
                        n_subjects=n_subjects,
                        n_observations=n_observations,
                        n_subjects_contrast=_paired_subject_count_for_trial(muscle_df, trial),
                        control_mean_ratio=_mean_ratio_for_cell(muscle_df, "control", trial),
                        vibration_mean_ratio=_mean_ratio_for_cell(muscle_df, "vibration", trial),
                    )
                )
            continue

        formula = f"{PRIMARY_VAR} ~ 0 + C(cell)"

        try:
            result, model_status, warning_text = _fit_mixedlm_with_fallback(
                formula=formula,
                data=muscle_df,
                groups_col="subject_id",
            )
        except Exception as exc:
            for trial in INTERGROUP_TRIALS:
                rows.append(
                    _empty_intergroup_lmm_row(
                        dataset_name=dataset_name,
                        muscle=muscle,
                        trial=trial,
                        model_status=f"model_error: {exc}",
                        n_subjects=n_subjects,
                        n_observations=n_observations,
                        n_subjects_contrast=_paired_subject_count_for_trial(muscle_df, trial),
                        control_mean_ratio=_mean_ratio_for_cell(muscle_df, "control", trial),
                        vibration_mean_ratio=_mean_ratio_for_cell(muscle_df, "vibration", trial),
                    )
                )
            continue

        fe_params = result.fe_params
        param_names = list(fe_params.index)

        try:
            cov_fe = result.cov_params().loc[param_names, param_names]
        except Exception:
            cov_fe = None

        for trial in INTERGROUP_TRIALS:
            control_cell = f"control_{trial}"
            vibration_cell = f"vibration_{trial}"
            control_term = f"C(cell)[{control_cell}]"
            vibration_term = f"C(cell)[{vibration_cell}]"

            control_n = int((muscle_df["cell"] == control_cell).sum())
            vibration_n = int((muscle_df["cell"] == vibration_cell).sum())
            n_subjects_contrast = _paired_subject_count_for_trial(muscle_df, trial)

            control_mean_ratio = _mean_ratio_for_cell(muscle_df, "control", trial)
            vibration_mean_ratio = _mean_ratio_for_cell(muscle_df, "vibration", trial)

            if control_term not in param_names or vibration_term not in param_names:
                rows.append(
                    _empty_intergroup_lmm_row(
                        dataset_name=dataset_name,
                        muscle=muscle,
                        trial=trial,
                        model_status="contrast_not_estimable_missing_cell",
                        n_subjects=n_subjects,
                        n_observations=n_observations,
                        n_subjects_contrast=n_subjects_contrast,
                        control_mean_ratio=control_mean_ratio,
                        vibration_mean_ratio=vibration_mean_ratio,
                    )
                )
                continue

            if control_n < 3 or vibration_n < 3:
                rows.append(
                    _empty_intergroup_lmm_row(
                        dataset_name=dataset_name,
                        muscle=muscle,
                        trial=trial,
                        model_status="contrast_not_estimable_not_enough_observations",
                        n_subjects=n_subjects,
                        n_observations=n_observations,
                        n_subjects_contrast=n_subjects_contrast,
                        control_mean_ratio=control_mean_ratio,
                        vibration_mean_ratio=vibration_mean_ratio,
                    )
                )
                continue

            if cov_fe is None:
                rows.append(
                    _empty_intergroup_lmm_row(
                        dataset_name=dataset_name,
                        muscle=muscle,
                        trial=trial,
                        model_status="contrast_not_estimable_no_covariance_matrix",
                        n_subjects=n_subjects,
                        n_observations=n_observations,
                        n_subjects_contrast=n_subjects_contrast,
                        control_mean_ratio=control_mean_ratio,
                        vibration_mean_ratio=vibration_mean_ratio,
                    )
                )
                continue

            contrast_vector = pd.Series(0.0, index=param_names)
            contrast_vector.loc[vibration_term] = 1.0
            contrast_vector.loc[control_term] = -1.0

            estimate = float(np.dot(contrast_vector.values, fe_params.loc[param_names].values))
            variance = float(
                np.dot(
                    contrast_vector.values,
                    np.dot(cov_fe.values, contrast_vector.values),
                )
            )

            if variance < 0 and np.isclose(variance, 0):
                variance = 0.0

            standard_error = np.sqrt(variance) if variance >= 0 else np.nan

            if pd.notna(standard_error) and standard_error > 0:
                z_value = estimate / standard_error
                p_raw = float(2 * stats.norm.sf(abs(z_value)))
                ci_low_log = estimate - z_crit * standard_error
                ci_high_log = estimate + z_crit * standard_error
            else:
                z_value = np.nan
                p_raw = np.nan
                ci_low_log = np.nan
                ci_high_log = np.nan

            ratio_of_ratios = np.exp(estimate)
            percent_difference = (ratio_of_ratios - 1) * 100

            rows.append(
                {
                    "dataset": dataset_name,
                    "muscle": MUSCLE_LABELS.get(muscle, muscle),
                    "trial": trial,
                    "time_point": trial,
                    "analysis_type": "intergroup_lmm",
                    "model_name": model_name,
                    "comparison": f"vibration_{trial} - control_{trial}",
                    "contrast": f"{vibration_term} - {control_term}",
                    "estimate_log_difference": clean_float(estimate),
                    "mean_log_difference_vibration_minus_control": clean_float(estimate),
                    "standard_error": clean_float(standard_error),
                    "std_error": clean_float(standard_error),
                    "z_value": clean_float(z_value),
                    "p_raw": p_raw,
                    "p_value": p_raw,
                    "p_adj": np.nan,
                    "p_value_bh": np.nan,
                    "significant_raw": bool(pd.notna(p_raw) and p_raw < 0.05),
                    "significant_adj": False,
                    "significant_bh": False,
                    "ratio_of_ratios": clean_float(ratio_of_ratios),
                    "percent_difference_vibration_vs_control": clean_float(percent_difference),
                    "ci_low_log": clean_float(ci_low_log),
                    "ci_high_log": clean_float(ci_high_log),
                    "ci_low_ratio": clean_float(np.exp(ci_low_log)),
                    "ci_high_ratio": clean_float(np.exp(ci_high_log)),
                    "control_mean_mnf_ratio": control_mean_ratio,
                    "vibration_mean_mnf_ratio": vibration_mean_ratio,
                    "n_subjects": int(n_subjects),
                    "n_observations": int(n_observations),
                    "n_subjects_contrast": int(n_subjects_contrast),
                    "model_status": model_status,
                    "convergence_status": model_status,
                }
            )

    results = pd.DataFrame(rows)

    if not results.empty:
        results = apply_bh_correction(
            results,
            p_col="p_raw",
            group_cols=["dataset", "muscle", "analysis_type"],
        )

        results["p_adj"] = results["p_value_bh"]
        results["significant_adj"] = results["significant_bh"]
        results["p_value"] = results["p_raw"]

    return results

def summarize_qc(df):
    qc_counts = (
        df
        .groupby(["qc_flag"], dropna=False)
        .size()
        .reset_index(name="n_rows")
        .sort_values("n_rows", ascending=False)
    )

    qc_by_subject = (
        df
        .groupby(["subject_id", "session_norm", "muscle_norm", "qc_flag"], dropna=False)
        .size()
        .reset_index(name="n_rows")
        .sort_values(["subject_id", "session_norm", "muscle_norm", "qc_flag"])
    )

    return qc_counts, qc_by_subject


def plot_mnf_ratio(df, dataset_name):
    PLOTS_DIR.mkdir(parents=True, exist_ok=True)

    session_colors = {
        "control": "#2ca02c",
        "vibration": "#ff7f0e",
    }

    rng = np.random.default_rng(12345)

    for muscle in MUSCLE_ORDER:
        subset = df[
            (df["muscle_norm"] == muscle)
            & (df[PLOT_VAR].notna())
        ].copy()

        if subset.empty:
            continue

        fig, ax = plt.subplots(figsize=(10, 5.5))

        x_base = np.arange(len(TRIAL_ORDER))
        box_width = 0.32

        offsets = {
            "control": -0.18,
            "vibration": 0.18,
        }

        for session in SESSION_ORDER:
            session_data = []
            session_positions = []
            mean_positions = []
            mean_values = []

            for i, trial in enumerate(TRIAL_ORDER):
                values = subset.loc[
                    (subset["session_norm"] == session)
                    & (subset["trial_cat"] == trial),
                    PLOT_VAR,
                ].dropna().to_numpy(dtype=float)

                if len(values) == 0:
                    continue

                position = x_base[i] + offsets[session]

                session_data.append(values)
                session_positions.append(position)

                mean_positions.append(position)
                mean_values.append(np.mean(values))

            if not session_data:
                continue

            color = session_colors.get(session, "gray")

            box = ax.boxplot(
                session_data,
                positions=session_positions,
                widths=box_width,
                patch_artist=True,
                showfliers=False,
                medianprops={"color": "black", "linewidth": 1.2},
                boxprops={"linewidth": 1.2},
                whiskerprops={"linewidth": 1.0},
                capprops={"linewidth": 1.0},
            )

            for patch in box["boxes"]:
                patch.set_facecolor(color)
                patch.set_alpha(0.45)
                patch.set_edgecolor(color)

            for whisker in box["whiskers"]:
                whisker.set_color(color)

            for cap in box["caps"]:
                cap.set_color(color)

            # Individual data points
            for pos, values in zip(session_positions, session_data):
                jitter = rng.normal(loc=0, scale=0.025, size=len(values))
                ax.scatter(
                    np.full(len(values), pos) + jitter,
                    values,
                    s=20,
                    alpha=0.55,
                    color=color,
                    edgecolors="none",
                    zorder=3,
                )

            # Group mean connected over time
            ax.plot(
                mean_positions,
                mean_values,
                marker="o",
                linestyle="--",
                linewidth=1.8,
                markersize=5,
                color=color,
                label=SESSION_LABELS.get(session, session),
            )

        ax.axhline(
            y=1.0,
            linestyle="--",
            linewidth=1,
            color="gray",
            alpha=0.9,
        )

        y_min = subset[PLOT_VAR].min()
        y_max = subset[PLOT_VAR].max()

        if pd.notna(y_min) and pd.notna(y_max):
            padding = (y_max - y_min) * 0.12
            if padding == 0:
                padding = 0.1

            ax.set_ylim(
                max(0, y_min - padding),
                y_max + padding,
            )

        ax.set_xticks(x_base)
        ax.set_xticklabels(
            [TRIAL_LABELS[t] for t in TRIAL_ORDER],
            rotation=45,
        )

        ax.set_ylabel("MNF ratio relative to Pre")
        ax.set_xlabel("Follow-up point")
        ax.set_title(
            f"MNF ratio over time - {MUSCLE_PLOT_LABELS.get(muscle, muscle)}"
        )

        ax.grid(True, axis="y", alpha=0.3)
        ax.set_xlim(-0.6, len(TRIAL_ORDER) - 0.4)
        ax.legend(title="Session")

        fig.tight_layout()

        output_path = PLOTS_DIR / f"mnf_ratio_{muscle}_{dataset_name}.png"
        fig.savefig(output_path, dpi=300, bbox_inches="tight")
        plt.close(fig)

def format_excel_file(excel_path):
    workbook = load_workbook(excel_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        if worksheet.max_row > 1 and worksheet.max_column > 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)

            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max(max_length + 2, 10), 35)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(excel_path)


def make_readme(main_df):
    rows = [
        {
            "section": "Input",
            "description": "File analysed",
            "value": str(INPUT_FILE),
        },
        {
            "section": "Primary variable",
            "description": "Variable used for statistical models",
            "value": PRIMARY_VAR,
        },
        {
            "section": "Plot variable",
            "description": "Variable used for figures",
            "value": PLOT_VAR,
        },
        {
            "section": "Main dataset",
            "description": "Rows included",
            "value": len(main_df),
        },
        {
            "section": "Intragroup analysis",
            "description": "Question answered",
            "value": "Within each session and muscle, each trial was compared against Pre using a linear mixed model.",
        },
        {
            "section": "Intergroup analysis",
            "description": "Question answered",
            "value": "For each post-intervention trial and muscle, vibration was compared against control using a linear mixed model on log-ratios with subject_id as a random effect.",
        },
        {
            "section": "Multiple comparisons",
            "description": "Correction used",
            "value": "Benjamini-Hochberg FDR correction within each analysis block.",
        },
        {
            "section": "Interpretation",
            "description": "MnF ratio",
            "value": "A ratio of 1 indicates no change from Pre; values below 1 indicate lower MnF than Pre; values above 1 indicate higher MnF than Pre.",
        },
    ]

    return pd.DataFrame(rows)


def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    PLOTS_DIR.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(INPUT_FILE, sheet_name=INPUT_SHEET)
    df = prepare_data(df)

    main_all_valid = df[
        (df["feature_status_norm"] == "ok")
        & (df["include_analysis_bool"])
        & (df[PRIMARY_VAR].notna())
        & np.isfinite(df[PRIMARY_VAR])
    ].copy()

    print(f"Rows in main analysis: {len(main_all_valid)}")

    descriptive_main = summarize_mnf(
        main_all_valid,
        dataset_name="main_all_valid",
    )

    intragroup_main, notes_main = run_intragroup_lmm(
        main_all_valid,
        dataset_name="main_all_valid",
    )

    intergroup_main = run_intergroup_tests(
        main_all_valid,
        dataset_name="main_all_valid",
    )

    qc_counts, qc_by_subject = summarize_qc(df)

    model_notes = notes_main.copy()

    if model_notes.empty:
        model_notes = pd.DataFrame(
            [
                {
                    "dataset": "",
                    "analysis": "",
                    "muscle": "",
                    "session": "",
                    "note": "No model warnings or errors.",
                }
            ]
        )
    
    plot_mnf_ratio(main_all_valid, dataset_name="main_all_valid")


    significant_tables = []

    if not intragroup_main.empty and "significant_bh" in intragroup_main.columns:
        temp = intragroup_main[intragroup_main["significant_bh"]].copy()
        temp["source_table"] = "intragroup_lmm"
        significant_tables.append(temp)

    if not intergroup_main.empty and "significant_bh" in intergroup_main.columns:
        temp = intergroup_main[intergroup_main["significant_bh"]].copy()
        temp["source_table"] = "intergroup_lmm"
        significant_tables.append(temp)

    if significant_tables:
        significant_results = pd.concat(significant_tables, ignore_index=True, sort=False)
    else:
        significant_results = pd.DataFrame()

    analysis_info = make_readme(main_all_valid)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        analysis_info.to_excel(writer, sheet_name="analysis_info", index=False)
        descriptive_main.to_excel(writer, sheet_name="descriptives", index=False)
        intragroup_main.to_excel(writer, sheet_name="intragroup_lmm", index=False)
        intergroup_main.to_excel(writer, sheet_name="intergroup_tests", index=False)
        significant_results.to_excel(writer, sheet_name="significant_results", index=False)

        if not model_notes.empty and not (
            len(model_notes) == 1
            and str(model_notes.iloc[0].get("note", "")) == "No model warnings or errors."
        ):
            model_notes.to_excel(writer, sheet_name="model_notes", index=False)

    format_excel_file(OUTPUT_FILE)

    print("\nFrequency statistics finished.")
    print(f"Results saved in:\n{OUTPUT_FILE}")
    print(f"Plots saved in:\n{PLOTS_DIR}")


if __name__ == "__main__":
    main()
